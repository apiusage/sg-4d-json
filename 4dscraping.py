import os
import requests
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font

# -------------------------
# CONFIG
# -------------------------
JSON_URL = "https://raw.githubusercontent.com/apiusage/sg-4d-json/main/4d.json"
EXCEL_FILE = "4d_history_results.xlsx"
SHEET_NAME = "4D"
SGT = ZoneInfo("Asia/Singapore")

# -------------------------
# FETCH NUMBERS
# -------------------------
def fetch_nums():
    r = requests.get(JSON_URL, timeout=10)
    r.raise_for_status()
    nums = r.json()
    return [str(n).zfill(4) for n in nums]

# -------------------------
# INIT EXCEL IF NOT EXISTS
# -------------------------
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME

        headers = ["DrawDate", "Day", "1st", "2nd", "3rd", "Starter", "Consolation"]
        ws.append(headers)

        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="FFFF00")
            cell.font = Font(bold=True)

        # Column widths
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 6
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 8
        ws.column_dimensions["E"].width = 8
        ws.column_dimensions["F"].width = 55
        ws.column_dimensions["G"].width = 55

        wb.save(EXCEL_FILE)
        print(f"📄 Created: {EXCEL_FILE}")


# -------------------------
# FETCH AND SAVE
# -------------------------
def fetch_and_save():
    try:
        nums = fetch_nums()

        if not nums or len(nums) < 23:
            print(f"❌ Insufficient data: got {len(nums)} numbers, need 23")
            return

        now = datetime.now(SGT)
        draw_date = now.strftime("%Y-%m-%d")
        day = now.strftime("%a")

        first  = nums[0]
        second = nums[1]
        third  = nums[2]
        starter     = ",".join(nums[3:13])   # 10 numbers
        consolation = ",".join(nums[13:23])  # 10 numbers

        # Init file if needed
        init_excel()

        wb = load_workbook(EXCEL_FILE)
        ws = wb[SHEET_NAME]

        # Duplicate check
        existing_dates = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        if draw_date in existing_dates:
            print(f"⚠️  {draw_date} already exists, skipping.")
            wb.close()
            return

        # Append row
        ws.append([draw_date, day, first, second, third, starter, consolation])

        # Force text format on number columns to prevent Excel mangling
        last_row = ws.max_row
        for col_idx in range(3, 8):  # C to G: 1st, 2nd, 3rd, Starter, Consolation
            cell = ws.cell(row=last_row, column=col_idx)
            cell.number_format = "@"

        wb.save(EXCEL_FILE)
        wb.close()

        print(f"✅ Saved {draw_date} ({day})")
        print(f"   1st: {first} | 2nd: {second} | 3rd: {third}")
        print(f"   Starter:     {starter}")
        print(f"   Consolation: {consolation}")

    except requests.RequestException as e:
        print(f"❌ Network error: {e}")
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()

# -------------------------
# MAIN
# -------------------------
if __name__ == "__main__":
    fetch_and_save()