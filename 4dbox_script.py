import os
import itertools
import requests
import pandas as pd
from collections import Counter, defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

URL = "https://raw.githubusercontent.com/apiusage/sg-4d-json/main/4d.json"
FILE = "4d_box_output.xlsx"

def fetch_numbers():
    try:
        return [str(n).zfill(4) for n in requests.get(URL, timeout=10).json()]
    except:
        return []

# ---------------------------
# Helper: ensure digits 0-9 present
# ---------------------------
def enforce_all_digits(box):
    """
    Ensure digits 0-9 appear at least once in a 4x4 box.
    Accepts box elements as either strings ('0'..'9') or integers (0..9).
    Returns same structure (4x4) with element types preserved.
    """
    # determine type of elements
    if not box or not box[0]:
        return box

    # flatten and maintain original element type (str or int)
    is_int = isinstance(box[0][0], int)
    flat = [str(d) for row in box for d in row]  # use strings for processing
    missing = sorted(set("0123456789") - set(flat))
    if not missing:
        # nothing missing, return original box
        return box

    # Replace duplicates with missing digits (replace first duplicate occurrences)
    # Prefer replacing elements that appear more than once
    for m in missing:
        replaced = False
        # Try to replace a duplicate digit
        for idx, d in enumerate(flat):
            if flat.count(d) > 1:
                flat[idx] = m
                replaced = True
                break
        # If no duplicates found (rare), replace any position that doesn't break column constraints
        if not replaced:
            for idx, d in enumerate(flat):
                if d not in missing:
                    flat[idx] = m
                    break

    # reshape back to 4x4
    fixed = []
    for i in range(0, 16, 4):
        row = flat[i:i+4]
        if is_int:
            row = [int(x) for x in row]
        fixed.append(row)
    return fixed

# ---------------------------
# Your original functions (with minimal edits)
# ---------------------------

def append_4d_results(csv="4d_results.csv"):
    nums = fetch_numbers()[:3]
    if len(nums) < 3:
        return None

    now = datetime.now()
    not_used = "_".join(sorted(set("0123456789") - set("".join(nums)))) + "_"
    row = {"DrawDate": now.strftime("%d/%m/%Y"), "1st": nums[0], "2nd": nums[1],
           "3rd": nums[2], "Days": now.strftime("%a"), "Not Used": not_used, "Year": now.year}

    df = pd.read_csv(csv) if os.path.exists(csv) else pd.DataFrame()
    if row["DrawDate"] not in df.get("DrawDate", []).values:
        pd.concat([df, pd.DataFrame([row])], ignore_index=True).to_csv(csv, index=False)
        print("Appended:", row)
    return row

def generate_box(numbers):
    numbers = [str(n).zfill(4) for n in numbers]
    counters = [Counter(num[i] for num in numbers) for i in range(4)]
    tops = [[d for d, _ in c.most_common(10)] for c in counters]
    box = [[tops[c][r] for c in range(4)] for r in range(4)]

    # Fix column duplicates and track digit columns
    digit_cols = defaultdict(set)
    for c in range(4):
        seen = set()
        for r in range(4):
            if box[r][c] in seen:
                box[r][c] = next(d for d in tops[c] if d not in seen)
            seen.add(box[r][c])
            digit_cols[box[r][c]].add(c)

    # Fix digits in >2 columns
    for digit, cols in list(digit_cols.items()):
        while len(cols) > 2:
            _, r, c = min((counters[c][digit], r, c)
                          for c in cols for r in range(4) if box[r][c] == digit)
            col_used = {box[i][c] for i in range(4)}
            box[r][c] = next(d for d in tops[c] if d not in col_used and len(digit_cols[d]) < 2)
            digit_cols[digit].remove(c)
            digit_cols[box[r][c]].add(c)

    # Fill missing digits
    flat = [d for row in box for d in row]
    for m in sorted(set('0123456789') - set(flat)):
        for i in range(15, -1, -1):
            r, c = divmod(i, 4)
            if flat.count(box[r][c]) > 1:
                box[r][c] = flat[i] = m
                break

    return box

def gen_perms(flat):
    res = []
    for i, j, k, l in itertools.product(range(0, 16, 4), range(1, 16, 4),
                                        range(2, 16, 4), range(3, 16, 4)):
        g = [flat[i], flat[j], flat[k], flat[l]]
        res.extend(''.join(p) for p in itertools.permutations(g))
    return res

def autofit(ws):
    for col in ws.columns:
        max_len = max((max((len(str(line)) for line in str(c.value or "").splitlines()), default=0) for c in col),
                      default=0)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max_len * 1.3 + 4, 120) if col_letter != 'A' else min(max_len + 2,
                                                                                                           20)

    for r in range(1, ws.max_row + 1):
        max_lines = max((len(str(ws.cell(r, c).value or "").splitlines()) for c in range(1, ws.max_column + 1)),
                        default=1)
        ws.row_dimensions[r].height = max_lines * 15

def run_4d_box(sheet="Perfect_4DBox"):
    numbers = fetch_numbers()
    box = generate_box(numbers)

    # Enforce presence of all digits 0-9 (preserve string type used elsewhere)
    box = enforce_all_digits(box)

    print("Generated 4x4 Box:\n" + '\n'.join(' '.join(row) for row in box))

    flat = [d for row in box for d in row]
    perms = gen_perms(flat)
    dedup = len(set(''.join(sorted(p)) for p in perms))
    unique = len(set(perms))
    matched = [n for n in numbers if n in perms]

    wb = load_workbook(FILE) if os.path.exists(FILE) else Workbook()
    ws = wb[sheet] if sheet in wb.sheetnames else wb.create_sheet(sheet)

    ws.insert_rows(2)
    ws['A2'] = datetime.today().strftime("%d/%m/%Y (%a)")
    ws['B2'] = '\n'.join(' '.join(row) for row in box)
    ws['C2'] = f"✅ Direct: {len(matched)}/{unique} ({len(matched) / unique * 100:.2f}%)\n✅ iBet: {len(matched)}/{dedup} ({len(matched) / dedup * 100:.2f}%)"
    ws['B2'].alignment = ws['C2'].alignment = Alignment(wrapText=True)
    ws['B2'].font = Font(name="Courier New")

    autofit(ws)
    wb.save(FILE)
    print(f"✅ 4D box saved to '{sheet}'")

def update_4d_box_stats(sheet="Predicted_Box"):
    past = fetch_numbers()
    if not os.path.exists(FILE):
        return

    wb = load_workbook(FILE)
    if sheet not in wb.sheetnames:
        return

    ws = wb[sheet]
    for r in range(2, ws.max_row + 1):
        b, s = ws[f"B{r}"].value, ws[f"C{r}"]
        if not b or s.value:
            continue

        flat = [d for d in b if d.isdigit()]
        perms = gen_perms(flat)
        dedup = len(set(''.join(sorted(p)) for p in perms))
        unique = len(set(perms))
        direct = sum(1 for n in past if n in perms)
        ibet = len({n for n in past if ''.join(sorted(n)) in {''.join(sorted(p)) for p in perms}})

        s.value = f"▶ iBet: {ibet}/{dedup} ({ibet / dedup * 100:.2f}%)\n▶ Direct: {direct}/{unique} ({direct / unique * 100:.2f}%)\n▶ Total Sets hit: {direct}"
        s.alignment = Alignment(wrapText=True)

    autofit(ws)
    wb.save(FILE)
    print(f"✅ Stats updated in '{sheet}'")

def generate_predicted_box(sheet_in="Perfect_4DBox", sheet_out="Predicted_Box"):
    try:
        df = pd.read_excel(FILE, sheet_name=sheet_in) if os.path.exists(FILE) else \
            pd.read_excel(
                BytesIO(requests.get(f"https://github.com/apiusage/sg-4d-json/raw/main/{FILE}", timeout=10).content),
                sheet_name=sheet_in)
    except:
        return

    boxes = [[int(d) for d in ''.join(filter(str.isdigit, str(b)))]
             for b in df.iloc[:, 1].dropna() if any(c.isdigit() for c in str(b))]

    pos_counts = [defaultdict(int) for _ in range(16)]
    for box in boxes:
        for i, d in enumerate(box[:16]):
            pos_counts[i][d] += 1

    pos_probs = [{d: c / sum(pos.values()) for d, c in pos.items()} if sum(pos.values()) > 0 else {}
                 for pos in pos_counts]

    def pick(p, r, c, col_cnt):
        avail = {d: v for d, v in p.items() if d not in r and d not in c and col_cnt.get(d, 0) < 2}
        if avail:
            return max(avail, key=avail.get)
        return next((d for d in range(10) if d not in r and d not in c and col_cnt.get(d, 0) < 2), 0)

    box, col_digits, col_cnt = [[-1] * 4 for _ in range(4)], [[] for _ in range(4)], {}
    for r in range(4):
        for c in range(4):
            d = pick(pos_probs[r * 4 + c], box[r], col_digits[c], col_cnt)
            box[r][c] = d
            col_digits[c].append(d)
            col_cnt[d] = sum(1 for col in col_digits if d in col)

    # enforce all digits appear at least once (generate_predicted_box uses integers)
    box = enforce_all_digits(box)

    print("Predicted Box:\n" + '\n'.join(' '.join(str(x) for x in row) for row in box))

    sg_tz = ZoneInfo("Asia/Singapore")
    next_draw = datetime.now(sg_tz) + timedelta(days=1)
    while next_draw.weekday() not in [2, 5, 6]:
        next_draw += timedelta(days=1)

    wb = load_workbook(FILE) if os.path.exists(FILE) else Workbook()
    ws = wb[sheet_out] if sheet_out in wb.sheetnames else wb.create_sheet(sheet_out)

    ws.insert_rows(2)
    ws['A2'] = next_draw.strftime("%d/%m/%Y (%a)")
    ws['B2'] = '\n'.join(' '.join(str(x) for x in row) for row in box)
    ws['B2'].alignment = Alignment(wrapText=True, vertical="top")
    ws['B2'].font = Font(name="Courier New")

    autofit(ws)
    wb.save(FILE)
    print(f"✅ Predicted box saved to '{sheet_out}'")

if __name__ == "__main__":
    append_4d_results()
    run_4d_box()
    update_4d_box_stats()
    generate_predicted_box()