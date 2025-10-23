import requests
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import os

# --- Fetch numbers ---
url = "https://raw.githubusercontent.com/apiusage/sg-4d-json/main/4d.json"
nums = requests.get(url).json()

# --- Prepare row ---
row = [
    datetime.now().strftime("%a (%Y-%m-%d)"),
    int(nums[0]), int(nums[1]), int(nums[2]),
    " ".join(nums[3:13]),
    " ".join(nums[13:23])
]

f = "4d_results_all.xlsx"
headers = ["DrawDate","1st","2nd","3rd","Starter","Consolation"]

if not os.path.exists(f):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(headers)
    fill = PatternFill("solid", fgColor="FFFF00")
    for c in ws[1]: c.fill, c.font = fill, Font(bold=True)
    wb.save(f)

wb = load_workbook(f)
ws = wb["Results"]

# --- Avoid duplicate DrawDate ---
dates = [cell.value for cell in ws["A"]]
if row[0] in dates:
    exit(print(f"⚠️ {row[0]} already exists."))

ws.append(row)
wb.save(f)
print(f"✅ Saved 4D results for {row[0]}.")
