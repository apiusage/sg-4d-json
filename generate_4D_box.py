import pandas as pd, os, requests
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime
from io import BytesIO

# ----------------- CONFIG -----------------
GITHUB_EXCEL_URL = "https://github.com/apiusage/sg-4d-json/raw/main/4d_box_output.xlsx"
DESIRED_SHEET_NAME = "Perfect_4DBox"   # <-- sheet to read from GitHub
OUTPUT_FILE = "4d_box_output.xlsx"
OUTPUT_SHEET = "Predicted_Box"          # <-- Save new predicted boxes here
# -----------------------------------------

# --- Load past boxes from GitHub Excel ---
resp = requests.get(GITHUB_EXCEL_URL, timeout=10)
resp.raise_for_status()
xls = pd.ExcelFile(BytesIO(resp.content))

# Clean sheet names to avoid hidden characters
sheet_names_clean = [name.strip() for name in xls.sheet_names]

if DESIRED_SHEET_NAME not in sheet_names_clean:
    print(f"⚠ Sheet '{DESIRED_SHEET_NAME}' not found. Script will exit without generating box.")
    exit()

# Exact sheet name match
idx = sheet_names_clean.index(DESIRED_SHEET_NAME)
PAST_SHEET_NAME = xls.sheet_names[idx]

df = pd.read_excel(BytesIO(resp.content), sheet_name=PAST_SHEET_NAME)
print(f"✅ Loaded sheet '{PAST_SHEET_NAME}' successfully. Total rows: {len(df)}")

# --- Flatten boxes into digits ---
col_values = df.iloc[:, 1].dropna().tolist()  # 2nd column
boxes = []
for box in col_values:
    digits = [int(d) for d in ''.join(filter(str.isdigit, str(box)))]
    if digits:
        boxes.append(digits)

print(f"Total past boxes read from '{PAST_SHEET_NAME}': {len(boxes)}")

# --- Compute position-wise frequencies ---
position_counts = [defaultdict(int) for _ in range(16)]
for box in boxes:
    for i, digit in enumerate(box[:16]):
        position_counts[i][digit] += 1

position_probs = []
for pos_dict in position_counts:
    total = sum(pos_dict.values())
    probs = {digit: count / total for digit, count in pos_dict.items()} if total > 0 else {}
    position_probs.append(probs)

# --- Deterministic digit selection ---
def select_digit(prob_dict, exclude_row, exclude_col, digit_column_count):
    available = {d: p for d, p in prob_dict.items()
                 if d not in exclude_row and d not in exclude_col and digit_column_count.get(d, 0) < 2}
    if not available:
        choices = [d for d in range(10)
                   if d not in exclude_row and d not in exclude_col and digit_column_count.get(d, 0) < 2]
        return choices[0] if choices else 0
    return max(available, key=available.get)

# --- Generate deterministic 4x4 box ---
def generate_box():
    box = [[-1] * 4 for _ in range(4)]
    col_digits_list = [[] for _ in range(4)]
    digit_column_count = {}

    for r in range(4):
        for c in range(4):
            pos = r * 4 + c
            exclude_row = box[r]
            exclude_col = col_digits_list[c]
            digit = select_digit(position_probs[pos], exclude_row, exclude_col, digit_column_count)
            box[r][c] = digit
            col_digits_list[c].append(digit)
            cols_with_digit = [i for i, col in enumerate(col_digits_list) if digit in col]
            digit_column_count[digit] = len(cols_with_digit)

    # ensure all digits 0–9 appear at least once
    flat_box = [d for row in box for d in row]
    missing_digits = [d for d in range(10) if d not in flat_box]
    for d in missing_digits:
        for r in range(4):
            for c in range(4):
                current_digit = box[r][c]
                if flat_box.count(current_digit) > 1 and digit_column_count.get(d, 0) < 2:
                    col_digits_list[c].remove(current_digit)
                    box[r][c] = d
                    col_digits_list[c].append(d)
                    flat_box = [x for row in box for x in row]
                    digit_column_count[d] = len(
                        [i for i, col in enumerate(col_digits_list) if d in col]
                    )
                    break
            if d in flat_box:
                break

    return box

# --- Generate box ---
predicted_box = generate_box()
print("Predicted 4x4 Box (unique columns, max 2 columns per digit):")
for row in predicted_box:
    print('  '.join(str(d) for d in row))  # 2 spaces per digit

# --- Prepare string for Excel ---
box_str = '\n'.join('  '.join(str(d) for d in row) for row in predicted_box) + '\n'
date_str = datetime.today().strftime("%d/%m/%Y (%a)")

# --- Save to Excel ---
wb = load_workbook(OUTPUT_FILE) if os.path.exists(OUTPUT_FILE) else Workbook()
ws = wb[OUTPUT_SHEET] if OUTPUT_SHEET in wb.sheetnames else wb.create_sheet(OUTPUT_SHEET)

# Insert new row at top
ws.insert_rows(2)
if ws.max_row == 1:
    ws['A1'], ws['B1'], ws['C1'] = "Date", "4x4 Box", "Stats"

# Write safe text (monospace font + wrap text)
ws['A2'] = date_str
ws['B2'] = box_str.strip()
ws['B2'].number_format = '@'  # Force plain text
ws['B2'].alignment = Alignment(wrapText=True, vertical="top")
ws['B2'].font = Font(name="Courier New")  # Monospace = perfect alignment on mobile
ws.column_dimensions['B'].width = 28

wb.save(OUTPUT_FILE)
print(f"✅ 4x4 box generated and saved to '{OUTPUT_FILE}' in sheet '{OUTPUT_SHEET}'.")
