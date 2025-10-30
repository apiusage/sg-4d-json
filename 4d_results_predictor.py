import os, random, pandas as pd, requests, math
from collections import Counter, defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

# Config
CSV_FILE, PRED_FILE = "4d_results_all.xlsx", "4d_box_output.xlsx"
RESULTS_SHEET, PRED_SHEET = "Results", "Direct_Prediction"
TOP_N, DECAY_DAYS = 5, 90.0
PRIZE_W = {"1st": 1.0, "2nd": 0.9, "3rd": 0.8, "Starter": 0.5, "Consolation": 0.3}
WEIGHTS = {"freq": 0.35, "pos": 0.3, "recency": 0.15, "balance": 0.1, "streak": 0.05, "rand": 0.05}
COLORS = ["CCFFFF", "CCFFCC", "FFCCCC", "FFFFCC"]
URL = "https://raw.githubusercontent.com/apiusage/sg-4d-json/main/4d.json"

def fetch_nums():
    return [str(n).zfill(4) for n in requests.get(URL, timeout=10).json()]

def next_draw():
    sg = ZoneInfo("Asia/Singapore")
    d = datetime.now(sg) + timedelta(days=1)
    while d.weekday() not in [2, 5, 6]:
        d += timedelta(days=1)
    return d.strftime("%d/%m/%Y (%a)")

def fetch_and_save_results():
    try:
        nums = fetch_nums()
        if not nums or len(nums) < 23:
            print("❌ Failed to fetch numbers or insufficient data")
            return

        now = datetime.now(ZoneInfo("Asia/Singapore"))
        row = [
            now.strftime("%a (%Y-%m-%d)"),
            nums[0],
            nums[1],
            nums[2],
            " ".join(str(n) for n in nums[3:13]),  # Starter prizes
            " ".join(str(n) for n in nums[13:23])  # Consolation prizes
        ]

        # Create file if it doesn't exist
        if not os.path.exists(CSV_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = RESULTS_SHEET
            ws.append(["DrawDate", "1st", "2nd", "3rd", "Starter", "Consolation"])
            for c in ws[1]:
                c.fill = PatternFill("solid", fgColor="FFFF00")
                c.font = Font(bold=True)
            wb.save(CSV_FILE)
            print(f"📄 Created new file: {CSV_FILE}")

        # Load and check for duplicates
        wb = load_workbook(CSV_FILE)
        ws = wb[RESULTS_SHEET]

        # Check existing dates (skip header row)
        existing_dates = [cell.value for cell in ws["A"][1:]]
        if row[0] in existing_dates:
            print(f"⚠️ {row[0]} already exists")
            wb.close()
            return

        # Append and save
        ws.append(row)
        wb.save(CSV_FILE)
        wb.close()
        print(f"✅ Saved results for {row[0]}: {row[1]}, {row[2]}, {row[3]}")

    except Exception as e:
        print(f"❌ Error in fetch_and_save_results: {e}")
        import traceback
        traceback.print_exc()

def calculate_stats(predicted):
    if not predicted:
        return "▶ iBet: 0/0 (0.00%)\n▶ Direct: 0/0 (0.00%)\n▶ Total Sets hit: 0"

    try:
        latest = fetch_nums()
        prize = {latest[0]: "1st", latest[1]: "2nd", latest[2]: "3rd"}
        for n in latest[3:13]:
            prize[n] = "Starter"
        for n in latest[13:23]:
            prize[n] = "Consolation"
    except Exception as e:
        return f"Error: {e}"

    direct, ibet = [], []
    for num in predicted:
        num = str(num).zfill(4)
        if num in prize:
            direct.append(f"{num} ({prize[num]})")
        else:
            for w, p in prize.items():
                if sorted(num) == sorted(w):
                    ibet.append(f"{num} → {w} ({p})")
                    break

    n, d, i = len(predicted), len(direct), len(ibet)
    return (f"▶ iBet: {i}/{n} ({i / n * 100:.2f}%) - {', '.join(ibet)}\n"
            f"▶ Direct: {d}/{n} ({d / n * 100:.2f}%) - {', '.join(direct)}\n"
            f"▶ Total Sets hit: {d}")

def ensure_headers(ws):
    headers = ["Date", "Predicted 4D Numbers", "Stats"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col)
        if not c.value:
            c.value = h
            c.fill = PatternFill("solid", fgColor="FFFF00")
            c.font = Font(bold=True)

def calculate_stats_in_excel():
    if not os.path.exists(PRED_FILE):
        return

    wb = load_workbook(PRED_FILE)
    ws = wb[PRED_SHEET] if PRED_SHEET in wb.sheetnames else wb.create_sheet(PRED_SHEET)
    ensure_headers(ws)

    if ws.max_row < 2:
        wb.save(PRED_FILE)
        return

    b = ws["B2"].value
    if not b or ws["C2"].value:
        wb.save(PRED_FILE)
        return

    stats = calculate_stats(b.split())
    ws["C2"].value = stats
    ws["C2"].alignment = Alignment(wrapText=True)

    last = next((i for i in range(ws.max_row, 1, -1) if ws.cell(i, 3).value), 1)
    ws["C2"].fill = PatternFill("solid", fgColor=COLORS[(last - 1) % len(COLORS)])
    ws["C2"].font = Font(bold=True)

    wb.save(PRED_FILE)
    print(f"✅ Stats calculated: {b}")

def predict_and_append_excel():
    df = pd.read_excel(CSV_FILE)
    records = []

    for _, row in df.iterrows():
        try:
            date = pd.to_datetime(str(row['DrawDate']).split('(')[1].split(')')[0])
        except:
            date = pd.to_datetime(str(row['DrawDate']), errors='coerce')
        if pd.isna(date):
            continue

        for t in ['1st', '2nd', '3rd']:
            if t in row and not pd.isna(row[t]):
                num = str(row[t]).strip().zfill(4)
                if num.isdigit() and len(num) == 4:
                    records.append((date, t, num))

        for col in ['Starter', 'Consolation']:
            if col in row and not pd.isna(row[col]):
                for n in str(row[col]).split():
                    num = n.strip().zfill(4)
                    if num.isdigit() and len(num) == 4:
                        records.append((date, col, num))

    records.sort(key=lambda x: x[0])
    if not records:
        return []

    full_freq, pos_freq = {}, [Counter() for _ in range(4)]
    recency, streaks = defaultdict(float), defaultdict(int)
    today = records[-1][0]

    for idx, (date, t, num) in enumerate(records):
        w = PRIZE_W.get(t, 0.5)
        full_freq[num] = full_freq.get(num, 0) + w
        for p, d in enumerate(num[:4]):
            pos_freq[p][d] += w
        recency[num] += w * (0.5 ** ((today - date).days / DECAY_DAYS))
        if idx > 0 and records[idx - 1][2] == num:
            streaks[num] += 1

    def norm(d):
        if not d:
            return {}
        mx, mn = max(d.values()), min(d.values())
        return {k: 1.0 for k in d} if mx == mn else {k: (v - mn) / (mx - mn) for k, v in d.items()}

    norm_full, norm_rec = norm(full_freq), norm(recency)
    max_streak = max(streaks.values()) if streaks else 1

    def pos_score(num):
        s = 1
        for p, d in enumerate(num):
            tot = sum(pos_freq[p].values())
            if tot:
                s *= pos_freq[p].get(d, 0) / tot
        return s

    def balance(num):
        digits = [int(d) for d in num]
        odd = sum(d % 2 for d in digits)
        high = sum(1 for d in digits if d >= 5)
        return (1 - abs(odd - (4 - odd)) / 4 + 1 - abs(high - (4 - high)) / 4) / 2

    candidates = set(full_freq.keys())
    combos = [[d] for d, _ in pos_freq[0].most_common(5)]
    for p in range(1, 4):
        combos = [c + [d] for c in combos for d, _ in pos_freq[p].most_common(5)]
    candidates.update(''.join(c) for c in combos)

    scores = {}
    for num in candidates:
        num = str(num).zfill(4)
        ps = pos_score(num)
        scores[num] = (WEIGHTS["freq"] * norm_full.get(num, 0) +
                       WEIGHTS["pos"] * (math.log1p(ps) if ps > 0 else 0) +
                       WEIGHTS["recency"] * norm_rec.get(num, 0) +
                       WEIGHTS["balance"] * balance(num) +
                       WEIGHTS["streak"] * (streaks.get(num, 0) / (1 + max_streak)) +
                       WEIGHTS["rand"] * random.uniform(0, 0.05))

    top = [n for n, _ in sorted(scores.items(), key=lambda x: x[1], reverse=True)[:TOP_N]]
    date_str = next_draw()

    wb = Workbook() if not os.path.exists(PRED_FILE) else load_workbook(PRED_FILE)
    ws = wb[PRED_SHEET] if PRED_SHEET in wb.sheetnames else wb.create_sheet(PRED_SHEET)
    ensure_headers(ws)

    ws.insert_rows(2)
    ws["A2"], ws["B2"], ws["C2"].value = date_str, ' '.join(top), ""

    last = next((i for i in range(ws.max_row, 1, -1) if ws.cell(i, 3).value), 1)
    ws["C2"].fill = PatternFill("solid", fgColor=COLORS[(last - 1) % len(COLORS)])
    ws["C2"].font = Font(bold=True)

    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = max((len(str(c.value or "")) for c in col),
                                                                           default=0) + 2
    wb.save(PRED_FILE)
    print(f"✅ Prediction for {date_str}")
    return top

if __name__ == "__main__":
    fetch_and_save_results()
    calculate_stats_in_excel()
    predicted = predict_and_append_excel()
    print("Predicted:", predicted)
