import pandas as pd, os, requests, itertools
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from io import BytesIO

# --- Fetch past winning numbers from GitHub ---
try:
    past_numbers = [str(n).zfill(4) for n in requests.get(
        "https://raw.githubusercontent.com/apiusage/sg-4d-json/refs/heads/main/4d.json", timeout=10
    ).json()]
    print(f"✅ Loaded {len(past_numbers)} past winning numbers")
except Exception as e:
    print(f"❌ Failed to fetch past numbers: {e}")
    past_numbers = []

# --- Load predicted boxes from Excel ---
fn = "4d_box_output.xlsx"
sheet_name = "Predicted_Box"

if not os.path.exists(fn):
    print(f"❌ Excel file '{fn}' not found")
    exit()

wb = load_workbook(fn)
if sheet_name not in wb.sheetnames:
    print(f"❌ Sheet '{sheet_name}' not found in '{fn}'")
    exit()

ws = wb[sheet_name]

# --- Process each row ---
for row_idx in range(2, ws.max_row + 1):
    box_cell = ws[f"B{row_idx}"]
    stats_cell = ws[f"C{row_idx}"]

    # Skip if no box or stats already filled
    if not box_cell.value or stats_cell.value:
        continue

    # Flatten the 4x4 box into digits
    flat_box = [d for d in box_cell.value if d.isdigit()]

    # --- Generate all iBet permutations ---
    perms = []
    for i, j, k, l in itertools.product(range(0,16,4), range(1,16,4), range(2,16,4), range(3,16,4)):
        if max(i, j, k, l) < 16:
            group = [flat_box[i], flat_box[j], flat_box[k], flat_box[l]]
            perms += [''.join(p) for p in itertools.permutations(group)]

    # --- Compute stats ---
    dedup_count = len(set(''.join(sorted(p)) for p in perms))  # iBet
    unique_count = len(set(perms))  # Direct
    direct_hits = sum(1 for n in past_numbers if n in perms)
    ibet_hits = len(set(n for n in past_numbers if ''.join(sorted(n)) in set(''.join(sorted(p)) for p in perms)))

    ibet_percent = ibet_hits / dedup_count * 100 if dedup_count else 0
    direct_percent = direct_hits / unique_count * 100 if unique_count else 0

    # --- Format stats ---
    stats_str = (
        f"▶ iBet Winning rate: {ibet_hits}/{dedup_count} ({ibet_percent:.2f}%)\n"
        f"▶ Direct Winning rate: {direct_hits}/{unique_count} ({direct_percent:.2f}%)\n"
        f"▶ Total Sets hit: {direct_hits}"
    )

    # Write to column C
    stats_cell.value = stats_str
    stats_cell.alignment = Alignment(wrapText=True)

# --- Save Excel ---
wb.save(fn)
print(f"✅ Stats updated in column C of '{fn}'")
